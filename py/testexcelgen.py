#!/usr/bin/python3.6
# -- coding: utf-8 --
import io
import sys
import os
import shutil
import requests
import pprint
import json
import pdb
import openpyxl
import cgi

APIURL = "https://houday.rbatos.com/api/api.php"
DOCDIR = "/var/www/html/docs"
TEMPLATEDIR = "/var/www/html/template2021"
ROOT = "/var/www/html/"
ONECELLARIAS = [
  'label', 'label0', 'label1', 'label2', 'label3', 
  'date', 'datetime', 'cname', 'scname', 'bname', 'sbname',
  'month', 'username', 'pname', 'comment', 'hno_user',
]

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def outPut(dt):
  print("Content-Type: application/json; charset=UTF-8\n\n")
  # print("Content-Type: text/html; charset=UTF-8;\n\n")
  print(dt)

def getDt():
  arg = cgi.FieldStorage()
  prms = {
    'a': 'fetchDocument',
    # 'hid': 'LE5MMsTF',
    # 'bid': 'p0CxjWNL',
    # 'stamp': '1612333843972',
    'hid': arg['hid'].value,
    'bid': arg['bid'].value,
    'stamp': arg['stamp'].value,
  }
  rt = requests.post(APIURL, prms)
  return(rt.text)

def handleFlile(dobj):
  # コピー先ファイル名を分解
  dobj['dst'] = os.path.split(dobj['dt'][0]['dst'])
  dstdir = DOCDIR + dobj['dst'][0]
  # ディレクトリ作成
  if os.path.exists(dstdir) == False:
    os.mkdir(dstdir)
  # ファイルコピー実施
  dstPath = DOCDIR + dobj['dt'][0]['dst']
  # アトリビュートごとコピーする パーミッションを引き継ぐため
  shutil.copy2(TEMPLATEDIR + '/' + dobj['dt'][0]['template'], dstPath)
  dobj['dstPath'] = dstPath
  return(dobj)

# 辞書配列の配列の中に要素があるか
def inDictList(key, array):
  r = False
  for a in array:
    if key in a:
      r = True
      break
  return(r)

# 辞書配列の配列の中の要素のインデックス
def indexDictList(key, array):
  r = -1
  i = 0
  for a in array:
    if key in a:
      r = i
      break
    i += 1
  return(r)

# セル範囲オブジェクトとデータオブジェクトで
# セル範囲に書き込みを行う
# セル範囲が単一か行単位、列単位、矩形を判断して処理する
# データが一次配列でセル範囲が行単位列単位ならそのままデータを流す
# データの形とセルの形が合わないと何が起こるか不明
def cellWriter(aria, data):
  # pdb.set_trace()
  if type(aria) is tuple:
    rows = len(aria)
    cols = len(aria[0])
  else:
    rows = 1
    cols = 1
  # 単一セルとして処理
  if rows == 1 and cols == 1:
    aria.value = data
  # 一列エリア
  elif rows > 1 and cols == 1:
    for i in range(len(data)):
      aria[i][0].value = data[i]
  # 矩形エリアと一行エリア
  else:
    for i in range(len(data)):
      # print('type(data[i])', type(data[i]))
      # print('data[i]', data[i])
      if type(data[i]) is list:
        for j in range(len(data[i])):
          aria[i][j].value = data[i][j]
      else:
        aria[0][i].value = data[i]

# handleExcelのマルチシート対応版
# dtのsheetsの中を読み取る
# 先頭のシートを基準にしてシートコピーしながら処理を進める
def handleExcelMS(fn, dt, hideEmpRows=False, proc=''):
  wb = openpyxl.load_workbook(fn)
  wsSrc = wb.worksheets[0] # コピー元となるシート
  # シート全体の名前付きエリアを取得
  # 名前付き範囲はシートコピーの影響を受けない
  arias = wb.defined_names.definedName
  locs = {} # 各エリア名に対応するロケーション文字列を格納する
  for aria in arias:
    locs[aria.name] = aria.attr_text.split('!')[1]
  # 1シートずつ処理実行
  for sData in dt['sheets']:
    trgWs = wb.copy_worksheet(wsSrc) # コピー実行
    trgWs.title = sData['sheetname']  # シート名変更
    # 格納済み名前付き範囲を舐める
    for loc in locs:
      # 定義済み名前が送信されてきたデータの中にあるか
      if loc in dt or loc in sData:
        if loc in dt:
          thisDt = dt[loc]
        if loc in sData:
          thisDt = sData[loc]
        thisAria = trgWs[locs[loc]]
        cellWriter(thisAria, thisDt)
  wb.remove(wsSrc)
  wb.save(fn)

# handleExcelのマルチシートシングルシート対応版
# dtのsheetsの中を読み取る
# 先頭のシートを基準にしてシートコピーしながら処理を進める
# ワークシートの範囲名とデータの要素名が一致したらデータを流し込む
def handleExcelMP(fn, dt, proc=''):
  wb = openpyxl.load_workbook(fn)
  wsSrc = wb.worksheets[0]  # コピー元となるシート
  # シート全体の名前付きエリアを取得
  # 名前付き範囲はシートコピーの影響を受けない
  arias = wb.defined_names.definedName
  locs = {}  # 各エリア名に対応するロケーション文字列を格納する
  for aria in arias:
    locs[aria.name] = aria.attr_text.split('!')[1]
  
  # 行削除を実施するか。行削除はエリア名に対して行われる
  # データに対してエリアが余っていたら削除実行
  # エリアが足りなかったらそもそも実行時エラーになる
  if 'hideEmpRows' in dt:
    hideEmpRows = dt['hideEmpRows']
  else:
    hideEmpRows = ''

  if 'sheets' in dt and len(dt['sheets']) >= 1:
    sDatas = dt['sheets']
  else:
    sDatas = [dt]  # sheetsがない場合は一個の配列を作る
  
  # 1シートずつ処理実行
  for sData in sDatas:
    if (len(sDatas) > 1):
      trgWs = wb.copy_worksheet(wsSrc)  # コピー実行
    else:
      trgWs = wsSrc # シングルシートはコピーせずにリネーム
    if 'sheetname' in sData:
      trgWs.title = sData['sheetname']  # シート名変更
    # 格納済み名前付き範囲を舐める
    for loc in locs:
      # 定義済み名前が送信されてきたデータの中にあるか
      if loc in dt or loc in sData:
        if loc in dt:
          thisDt = dt[loc]
        if loc in sData:
          thisDt = sData[loc]
        # pdb.set_trace()
        # print('locs[loc]', locs[loc], 'locs', locs, 'loc', loc)
        thisAria = trgWs[locs[loc]]
        cellWriter(thisAria, thisDt)
    # 行の非表示
    if hideEmpRows:
      delAria = trgWs[locs[hideEmpRows]]
      if hideEmpRows in sData:
        thisDt = sData[hideEmpRows]
      else:
        thisDt = dt[hideEmpRows]
      toprow = delAria[0][0].row  # 先頭行の取得 単一セルだとエラーになる
      ariaLength = len(delAria)
      dataLength = len(thisDt)
      # trgWs.delete_rows(toprow + dataLength, ariaLength - dataLength)
      for i in range(toprow + dataLength, ariaLength + toprow):
        trgWs.row_dimensions[i].hidden = True

  if (len(sDatas) > 1):
    wb.remove(wsSrc)
  wb.save(fn)


# シングルシート版作り直し
def handleExcelSS(fn, dt, hideEmpRows='', proc=''):
  wb = openpyxl.load_workbook(fn)
  ws = wb.worksheets[0] # 先頭のシートを設定
  # 名前付き範囲を取得
  arias = wb.defined_names.definedName
  locs = {}  # 各エリア名に対応するロケーション文字列を格納する
  for aria in arias:
    locs[aria.name] = aria.attr_text.split('!')[1]
  # 格納済み名前付き範囲を舐める
  for loc in locs:
    # 定義済み名前が送信されてきたデータの中にあるか
    if loc in dt:
      thisDt = dt[loc]
    thisAria = ws[locs[loc]]
    cellWriter(thisAria, thisDt)
  if 'sheetname' in dt:
    ws.title = dt['sheetname']
  wb.save(fn)

# fnのエクセルファイルの編集を行う
# dtは書き込みデータ
# procはプロセス名 あとから追加するかも
# hideEmpRowsでエリア内の不必要な行を非表示にする
def handleExcel(fn, dt, hideEmpRows=True, proc=''):
  # pdb.set_trace()
  wb = openpyxl.load_workbook(fn)
  aria = wb.defined_names['data']  # 名前付きのエリア取得
  ws = wb[aria.attr_text.split('!')[0]]
  cells = ws[aria.attr_text.split('!')[1]]
  topRow = cells[0][0].row  # エリアの先頭行
  ariaRows = len(cells)
  dtRows = len(dt['data'])
  i = 0
  for rowdt in dt['data']:
    j = 0
    for celldt in rowdt:
      cell = cells[i][j]
      # 空白行は処理しない
      if celldt != '':
        cell.value = celldt
      j += 1
    i += 1
  # 不必要な行を非表示にする
  if hideEmpRows:
    for i in range(topRow + dtRows, ariaRows + topRow):
      ws.row_dimensions[i].hidden = True
  for an in ONECELLARIAS:
    if an in wb.defined_names and an in dt:
      aria = wb.defined_names[an]
      ws = wb[aria.attr_text.split('!')[0]]
      cell = ws[aria.attr_text.split('!')[1]]
      # 範囲指定されている場合は左上をとる
      if type(cell) is tuple:
        cell = cell[0]
      if type(cell) is tuple:
        cell = cell[0]
      cell.value = dt[an]
  wb.save(fn)

if __name__ == "__main__":
  # dbより値取得
  rtn = {}
  # try:
  rtnStr = getDt()  # dbよりデータ取得
  dobj = json.loads(rtnStr)
  dobj = handleFlile(dobj) # ダウンロード用ディレクトリを作製してコピー実施
  rtn['handleFileDone'] = True
  dt = dobj['dt'][0]['content']
  # # 複数シート帳票
  # if len(dt['sheets']):
  #   handleExcelMS(dobj['dstPath'], dobj['dt'][0]['content'])
  # # 1シート帳票
  # else:
  #   handleExcelSS(dobj['dstPath'], dobj['dt'][0]['content'])
  # 汎用版
  handleExcelMP(dobj['dstPath'], dt)
  # else:
  #   raise Exception('データなし')
  
  rtn['handleExcelDone'] = True
  rtn['result'] = True
  rtn['dstPath'] = dobj['dt'][0]['dst']
  # rtn['root'] = ROOT
  # except Exception:
  #   rtn['result'] = False
  # else:
  #   rtn['finish'] = True

  outPut(json.dumps(rtn))
